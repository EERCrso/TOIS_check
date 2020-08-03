import openpyxl as opx
import datetime as dt
import numpy as np
from openpyxl.utils import get_column_letter


def get_TOIS_SLA(prio, code):
    """
    pozrie sa do excelu, v ktorom su SLA data pre TOIS
    :param code: L2 / L3 + O / V / R
    :param prio: priorita 1 / 2 / 3
    :return: hodnota a jednotka SLA casu
    """
    workbook_sla = opx.load_workbook(filename='SLA.xlsx')
    sheet_sla = workbook_sla.active

    # default A1 = empty
    row = 1
    col = 1
    val = np.nan
    unit = ""

    # najdi riadok hodnoty
    if code == 'L2O':
        row = 2
    elif code == 'L2R':
        row = 3
    elif code == 'L3O':
        row = 4
    elif code == 'L3V':
        row = 5
    elif code == 'L3R':
        row = 6

    # najdi stlpec hodnoty
    if prio == 1:
        col = 2
    elif prio == 2:
        col = 4
    elif prio == 3:
        col = 6

    val = sheet_sla.cell(row=row, column=col).value
    unit = sheet_sla.cell(row=row, column=col + 1).value

    # zavriet subor - read only
    workbook_sla.close()

    return val, unit


def excel_date(date1):
    """
    funkcia konvertuje casovy format z pythonu do excelu (len cele cislo)
    :param date1: datum+cas na konverziu v python formate datetime
    :return: skonvertovany datum+cas v excel formate
    """
    temp = dt.datetime(1899, 12, 30)  # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)


def auto_format_cell_width(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 1


def recognize_SLA_HPSM(str_sla):
    """
    :param str_sla: string, ktory je standardom pre HPSM automaticky generovane reporty
    v stlpci "SLT Name"
    :return: L2/L3 + O/R
    """
    if "odozvy" in str_sla:
        if "L3" in str_sla:
            return "L3O"
        else:
            return "L2O"
    else:  # riesenia
        if "L3" in str_sla:
            return "L3R"
        else:
            return "L2R"
