import pandas as pd
import aux_functions
import workdays as wd
from dateutil.relativedelta import *
import PySimpleGUI as sg
import openpyxl as opx
import datetime as dt
import numpy as np
from openpyxl.utils import get_column_letter
import pkg_resources.py2_warn


def get_TOIS_SLA(prio, code):
    """
    pozrie sa do excelu, v ktorom su SLA data pre TOIS
    :param code: L2 / L3 + O / V / R
    :param prio: priorita 1 / 2 / 3
    :return: hodnota a jednotka SLA casu
    """
    workbook_sla = opx.load_workbook(filename='SLA.xlsx')
    sheet_sla = workbook_sla.active

    # default A1 = empty
    row = 1
    col = 1
    val = np.nan
    unit = ""

    # najdi riadok hodnoty
    if code == 'L2O':
        row = 2
    elif code == 'L2R':
        row = 3
    elif code == 'L3O':
        row = 4
    elif code == 'L3V':
        row = 5
    elif code == 'L3R':
        row = 6

    # najdi stlpec hodnoty
    if prio == 1:
        col = 2
    elif prio == 2:
        col = 4
    elif prio == 3:
        col = 6

    val = sheet_sla.cell(row=row, column=col).value
    unit = sheet_sla.cell(row=row, column=col + 1).value

    # zavriet subor - read only
    workbook_sla.close()

    return val, unit


def excel_date(date1):
    """
    funkcia konvertuje casovy format z pythonu do excelu (len cele cislo)
    :param date1: datum+cas na konverziu v python formate datetime
    :return: skonvertovany datum+cas v excel formate
    """
    temp = dt.datetime(1899, 12, 30)  # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)


def auto_format_cell_width(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 1


def recognize_SLA_HPSM(str_sla):
    """
    :param str_sla: string, ktory je standardom pre HPSM automaticky generovane reporty
    v stlpci "SLT Name"
    :return: L2/L3 + O/R
    """
    if "odozvy" in str_sla:
        if "L3" in str_sla:
            return "L3O"
        else:
            return "L2O"
    else:  # riesenia
        if "L3" in str_sla:
            return "L3R"
        else:
            return "L2R"


# # 1.0 - NASTAVENIE PRISTUPOVYCH CIEST K SUBOROM + INE
#     # TODO - premysliet co bude lepsie - ci vyberat subor manualne, alebo prepisat v kode vzdy -> rychlejsie
#     csv_input_path = os.path.join('SKTTAC_29_07_2020.csv')  # export z TLN JIRA napr 01/06/2020
#     hpsm_input_path = os.path.join('Billien_202006_EMS.xlsx')  # data zo SKT za Billien za obdobie napr 05
#     report_path = os.path.join(
#         'TOIS_SLA_Report_2020-05_Interne_Vyhodnotenie_v4.xlsx')  # report z minuleho mesiaca (napr 04)
#     sht_name = 'TOIS všetky 05-2020'  # zvolit spravny sheet z minuleho mesiaca (napr 04)
#     hidden_cols = False
#     report_month = 6
#     report_year = 2020


def merge_exports(report_path, csv_input_path, hpsm_input_path, sht_name, report_month, report_year,
                  hidden_cols=False, to_gui=True):
    """
    hlavna funkcia ktora spaja vsetky reporty dokopy
    :param to_gui: vystup aj do GUI?
    :param report_path: report z minuleho mesiaca (napr 04)
    :param csv_input_path: export z TLN JIRA napr 01/06/2020
    :param hpsm_input_path: data zo SKT za Billien za obdobie napr 05
    :param sht_name: zvolit spravny sheet z minuleho mesiaca (napr 04)
    :param report_month: mesiac reportu, napr. 5
    :param report_year: rok reportu, napr. 2020
    :param hidden_cols: skryt niektore stplce v kompletnom exceli (nepovinne)
    :return: kratky a dlhy spojeny excel subor s incidentami do priecinka odkial sa spusta program
    """

    # 2.0 - EXPORT Z HPSM, UPRAVENY UZ CEZ VBA SKRIPT -> CITATELNEJSIA FORMA
    df_hpsm = pd.read_excel(hpsm_input_path)
    print('Subor ' + hpsm_input_path + ' uspesne nacitany ... ')
    if to_gui:
        window['MLINE'].update('Subor ' + hpsm_input_path + ' uspesne nacitany ... \n', append=True)

    # 2.1.0 - korekcia exportu z HPSM - ID v kazdom riadku spravne (nahradenie NaN hodnot)
    idx = np.nan
    row_counter = 0
    for ids in df_hpsm['Incident ID']:
        if ids is not np.nan:
            idx = ids
        elif ids is np.nan:
            df_hpsm.loc[row_counter, 'Incident ID'] = idx
        row_counter += 1
    del ids, idx, row_counter

    # 2.1.1 - korekcia porusenia SLA -> vzdy do prveho riadku z groupy ID pridat vsetky porusenia
    hpsm_breach_sla = ['L2 Odozva Breach HPSM', 'L2 Riesenie Breach HPSM',
                       'L3 Odozva Breach HPSM', 'L3 Riesenie Breach HPSM']
    for item in hpsm_breach_sla:
        df_hpsm.insert(loc=len(df_hpsm.columns), column=item, value='')
        df_breached_collected = pd.DataFrame(columns=df_hpsm.columns)

    for df_name, df_group in df_hpsm.groupby("Incident ID"):
        df_group_copy = df_group.copy()
        for index, values in df_group["SLT Name"].items():
            SLT_Breach = df_group.loc[index, 'SLT Breached']
            if recognize_SLA_HPSM(values) == 'L2O':
                df_group_copy.loc[:, 'L2 Odozva Breach HPSM'] = SLT_Breach
            elif recognize_SLA_HPSM(values) == 'L2R':
                df_group_copy.loc[:, 'L2 Riesenie Breach HPSM'] = SLT_Breach
            elif recognize_SLA_HPSM(values) == 'L3O':
                df_group_copy.loc[:, 'L3 Odozva Breach HPSM'] = SLT_Breach
            elif recognize_SLA_HPSM(values) == 'L3R':
                df_group_copy.loc[:, 'L3 Riesenie Breach HPSM'] = SLT_Breach

        df_breached_collected = df_breached_collected.append(df_group_copy, ignore_index=True)

    # 2.1.2 - pokial incident L3 a ma prazdne L2 polia -> default neporusenie SLA
    for index, values in df_breached_collected["L2 Odozva Breach HPSM"].items():
        if df_breached_collected.loc[index, 'L3 Odozva Breach HPSM'] and \
                df_breached_collected.loc[index, 'L3 Riesenie Breach HPSM']:
            df_breached_collected.loc[index, "L2 Odozva Breach HPSM"] = "Nie"

    for index, values in df_breached_collected["L2 Riesenie Breach HPSM"].items():
        if df_breached_collected.loc[index, 'L3 Odozva Breach HPSM'] and \
                df_breached_collected.loc[index, 'L3 Riesenie Breach HPSM']:
            df_breached_collected.loc[index, "L2 Riesenie Breach HPSM"] = "Nie"

    # TODO - DOPLNIT SLT BREACHED PRE VSETKY SLA CASY - AZ POTOM DELETE DUPLICATES
    # 2.2 - vyhodenie nepotrebnych stlpcov
    df_hpsm.drop('SLT Breached Next Month', axis=1, inplace=True)
    df_hpsm_new = df_breached_collected.copy()
    df_hpsm_new.drop(['SLT Name', 'SLT Expiration time', 'SLT Total time\nd hh:mi:ss', 'Open Time',
                      'Close Time', 'SLT Breached', 'SLT Breached Next Month'], axis=1, inplace=True)
    df_hpsm_new.drop_duplicates(subset='Incident ID', inplace=True, keep='first')

    # 2.2.1 - preindexovanie riadkov -> aby sa dal pouzit .loc neskor poporade v iteracii
    df_hpsm_new.reset_index(drop=True, inplace=True)

    # 2.3 - Definovanie novych stlpcov do prazdneho dataframe
    hpsm_cols_sla = ['L2 Odozva HPSM', 'L2 Riesenie HPSM',
                     'L3 Odozva HPSM', 'L3 Riesenie HPSM',
                     'L2 Odozva HPSM Total Time', 'L2 Riesenie HPSM Total Time',
                     'L3 Odozva HPSM Total Time', 'L3 Riesenie HPSM Total Time']
    for item in hpsm_cols_sla:
        df_hpsm_new.insert(loc=len(df_hpsm_new.columns), column=item, value='')

    # 2.4 - Definovanie hladanych velicin
    L2_odozva = np.nan
    L2_odozva_total = np.nan
    L2_riesenie = np.nan
    L2_riesenie_total = np.nan
    L3_odozva = np.nan
    L3_odozva_total = np.nan
    L3_riesenie = np.nan
    L3_riesenie_total = np.nan
    hpsm_problems = []

    # 2.5 - Hladanie velicin na zaklade charakteristickych popisov v stlpci SLT Name, pre kazde IM
    group_rows = 0
    for id_name, id_group in df_hpsm.groupby('Incident ID'):
        # kontrola na max 4 riadky v skupine, ak ano - vypise neskor ktore incidenty to boli
        if len(id_group.index) > 4:
            hpsm_problems.append(id_name)

        # preindexovanie na default id_group
        id_group.reset_index(drop=True, inplace=True)

        sla_rows = 0
        for sla in id_group['SLT Name']:
            if 'L3' not in sla:  # teda L2
                if 'Doba odozvy' in sla:
                    L2_odozva = id_group.loc[sla_rows, 'SLT Expiration time']
                    L2_odozva_total = id_group.loc[sla_rows, 'SLT Total time\nd hh:mi:ss']
                    df_hpsm_new.loc[group_rows, 'L2 Odozva HPSM'] = L2_odozva
                    df_hpsm_new.loc[group_rows, 'L2 Odozva HPSM Total Time'] = L2_odozva_total
                else:
                    L2_riesenie = id_group.loc[sla_rows, 'SLT Expiration time']
                    L2_riesenie_total = id_group.loc[sla_rows, 'SLT Total time\nd hh:mi:ss']
                    df_hpsm_new.loc[group_rows, 'L2 Riesenie HPSM'] = L2_riesenie
                    df_hpsm_new.loc[group_rows, 'L2 Riesenie HPSM Total Time'] = L2_riesenie_total
            else:  # teda L3
                if 'Doba odozvy' in sla:
                    L3_odozva = id_group.loc[sla_rows, 'SLT Expiration time']
                    L3_odozva_total = id_group.loc[sla_rows, 'SLT Total time\nd hh:mi:ss']
                    df_hpsm_new.loc[group_rows, 'L3 Odozva HPSM'] = L3_odozva
                    df_hpsm_new.loc[group_rows, 'L3 Odozva HPSM Total Time'] = L3_odozva_total
                else:
                    L3_riesenie = id_group.loc[sla_rows, 'SLT Expiration time']
                    L3_riesenie_total = id_group.loc[sla_rows, 'SLT Total time\nd hh:mi:ss']
                    df_hpsm_new.loc[group_rows, 'L3 Riesenie HPSM'] = L3_riesenie
                    df_hpsm_new.loc[group_rows, 'L3 Riesenie HPSM Total Time'] = L3_riesenie_total
            # potrebne pre zapis v groupach
            sla_rows += 1
        # potrebne pre zapis do finalneho dataframe
        group_rows += 1

    # ak bolo niekde viac hodnot ako bolo treba - vypisu sa
    if hpsm_problems:
        print("Problematicke incidenty v HPSM datach ( >4 riadky zaznamov): ")
        print(hpsm_problems)
        if to_gui:
            window['MLINE'].update("Problematicke incidenty v HPSM datach ( >4 riadky zaznamov): \n", append=True)
            window['MLINE'].update(hpsm_problems, append=True)
            window['MLINE'].update('\n', append=True)

    else:
        print('Vsetky data z HPSM uspesne konvertovane ...')
        if to_gui:
            window['MLINE'].update('Vsetky data z HPSM uspesne konvertovane ...\n', append=True)

    # premenovanie pre lepsiu orientaciu vo finalnom subore
    df_hpsm_new.rename(columns={'Title': 'Title - HPSM',
                                'Priority': 'Priority - HPSM',
                                'Status': 'Status - HPSM',
                                'L3 udrzba': 'L3 - HPSM',
                                'Is Outage': 'Outage - HPSM',
                                'SLT Start time': 'SLT Start time - HPSM',
                                }, inplace=True)

    # 3.0 - EXPORT -> TLN JIRA -> SKTTAC UPDATED IN 6 MONTHS -> EXPORT AS .CSV (delimiter = ;)
    used_cols = ['Issue id', 'Issue key', 'Summary', 'Issue Type',
                 'Status', 'Priority', 'Created', 'Updated', 'Resolved',
                 'Labels', 'Labels.1', 'Labels.2', 'Resolution', 'Assignee', 'Last Viewed',
                 'Description', 'Outward issue link (Duplicate)', 'Outward issue link (Relation )']
    custom_field_cols = ['Custom field (Ext ID)', 'Custom field (HPSM_Assigment_Group)',
                         'Custom field (HPSM_Assignee)', 'Custom field (HpsmIssueType)',
                         'Custom field (MEV ID)', 'Custom field (Module)',
                         'Custom field (Open - Closed)', 'Custom field (Open - Resolved)',
                         'Custom field (Reopening counter)', 'Custom field (Test Environment)']
    df_csv = pd.read_csv(csv_input_path, sep=';', decimal='.', usecols=used_cols + custom_field_cols)
    print('Zdrojovy JIRA subor uspesne nacitany ...')
    if to_gui:
        window['MLINE'].update('Zdrojovy JIRA subor uspesne nacitany ...\n', append=True)

    # premenovanie pre lepsiu orientaciu vo finalnom subore
    df_csv.rename(columns={'Issue id': 'Issue ID - JIRA',
                           'Issue key': 'Issue key - JIRA',
                           'Summary': 'Summary - JIRA',
                           'Issue Type': 'Issue Type - JIRA',
                           'Status': 'Status - JIRA',
                           'Priority': 'Priority - JIRA',
                           'Created': 'Created - JIRA',
                           'Updated': 'Updated - JIRA',
                           'Resolved': 'Resolved - JIRA',
                           'Labels': 'Label 1 - JIRA',
                           'Labels.1': 'Label 2 - JIRA',
                           'Labels.2': 'Label 3 - JIRA',
                           'Resolution': 'Resolution - JIRA',
                           'Assignee': 'Assignee - JIRA',
                           'Last Viewed': 'Last Viewed - JIRA',
                           'Description': 'Description - JIRA',
                           'Outward issue link (Duplicate)': 'Duplicate - JIRA',
                           'Outward issue link (Relation )': 'Relation to - JIRA',
                           'Custom field (Ext ID)': 'Custom field (Ext ID)',  # na tento stlpc je MERGE
                           'Custom field (HPSM_Assigment_Group)': 'HPSM Group - JIRA',
                           'Custom field (HPSM_Assignee)': 'Assignee HPSM - JIRA',
                           'Custom field (HpsmIssueType)': 'HPSM Issue Type - JIRA',
                           'Custom field (MEV ID)': 'MEV ID - JIRA',
                           'Custom field (Module)': 'Module - JIRA',
                           'Custom field (Open - Closed)': 'Open-Closed Time - JIRA',
                           'Custom field (Open - Resolved)': 'Open-Resolved Time - JIRA',
                           'Custom field (Reopening counter)': 'Reopen Counter - JIRA',
                           'Custom field (Test Environment)': 'Test Environment - JIRA'
                           }, inplace=True)

    # 3.1 - konverzia na datetime z csv formatu - pozor, natvrdo nastavene cez to_datetime!
    time_cols = ['Created - JIRA', 'Updated - JIRA', 'Resolved - JIRA', 'Last Viewed - JIRA']
    for dates in df_csv[time_cols]:
        df_csv[dates] = pd.to_datetime(df_csv[dates], format='%d.%m.%Y %H:%M', errors='ignore')

    # 3.2 - potrebne len Issue Type == Bug Ext, potom vymazat cely stlpec -> do buducna mozno obmedzit aj filtrom v JIRA
    df_csv.drop(df_csv[df_csv['Issue Type - JIRA'] != 'Bug Ext'].index, inplace=True)
    df_csv.drop('Issue Type - JIRA', axis=1, inplace=True)

    # 3.3 - pokial incident nema Ext ID -> pokusi sa nacitat zo Summary prvych 8 znakov
    for index, value in df_csv['Custom field (Ext ID)'].items():
        if value is np.nan:
            df_csv.loc[index, 'Custom field (Ext ID)'] = df_csv.loc[index, 'Summary - JIRA'][:8]

    # 3.4 - merge do spolocneho excelu JIRA + HPSM + report
    df_report = pd.read_excel(report_path, sheet_name=sht_name)
    df_merged_report = pd.merge(df_report, df_hpsm_new, how='outer',
                                left_on='Incident ID', right_on='Incident ID')

    # 3.4.1 - doplnenie novych incidentov -> Title z HPSM
    for index, value in df_merged_report['Title'].items():
        if value is np.nan:
            df_merged_report.loc[index, 'Title'] = df_merged_report.loc[index, 'Title - HPSM']

    # 3.4.2 - vymazanie duplicitnych stplcov Title
    df_hpsm_new.drop('Title - HPSM', axis=1, inplace=True)  # duplicita -> nachadza sa uz v df_report
    df_csv.drop('Summary - JIRA', axis=1, inplace=True)  # duplicita -> nachadza sa uz v df_report

    # 3.4.3 - finalny merge s JIRA exportom -> pokial prazdne polia, na incidente sa 6 mes+ nerobilo
    df_merged_all = pd.merge(df_merged_report, df_csv, how='left',
                             left_on='Incident ID', right_on='Custom field (Ext ID)')
    df_merged_all.drop(columns='Custom field (Ext ID)')  # duplikat po mergnuti

    # 3.5 - zoradenie stlpcov - lepsia kontrola neskor
    new_cols = ['Incident ID', 'Issue key - JIRA', 'Issue ID - JIRA',
                'Title',
                'Group', 'L3 - HPSM', 'Label 1 - JIRA', 'Label 2 - JIRA', 'Label 3 - JIRA',
                'P', 'Priority - HPSM', 'Priority - JIRA',
                'Status JIRA', 'Status - HPSM', 'Status - JIRA', 'Resolution - JIRA', 'Assignee - JIRA',
                'Assignee HPSM - JIRA',
                'Assign Time', 'Created - JIRA', 'SLT Start time - HPSM',
                'Čas parametra S.2', 'L2 Odozva HPSM', 'L2 Odozva HPSM Total Time', 'Splnenie parametra S.2',
                'L2 Odozva Breach HPSM',
                'Čas parametra S.3', 'L2 Riesenie HPSM', 'L2 Riesenie HPSM Total Time', 'Splnenie parametra S.3',
                'L2 Riesenie Breach HPSM',
                'Čas parametra S.4', 'L3 Odozva HPSM', 'L3 Odozva HPSM Total Time', 'Splnenie parametra S.4',
                'L3 Odozva Breach HPSM',
                'Čas parametra S.5', 'Splnenie parametra S.5',  # rovnake ako S.4 (L3 Odozva HPSM)
                'Čas parametra S.6', 'L3 Riesenie HPSM', 'L3 Riesenie HPSM Total Time', 'Splnenie parametra S.6',
                'L3 Riesenie Breach HPSM',
                'Updated - JIRA', 'Last Viewed - JIRA', 'Resolved - JIRA',
                'Description - JIRA',
                'Duplicate - JIRA', 'Relation to - JIRA', 'Outage - HPSM',
                'HPSM Group - JIRA', 'HPSM Issue Type - JIRA',
                'MEV ID - JIRA', 'Module - JIRA',
                'Open-Closed Time - JIRA', 'Open-Resolved Time - JIRA',
                'Reopen Counter - JIRA', 'Test Environment - JIRA']
    df_merged_all = df_merged_all[new_cols]

    # 3.6 - zoradenie na zaklade Incident ID zostupne
    df_merged_all.sort_values(by=['Incident ID'], ascending=False, ignore_index=True, inplace=True)

    # 3.7 - doplnenie casov, groupy a L3 novych (prazdnych) incidentov
    for index, value in df_merged_all['Group'].items():
        if value is np.nan:
            if df_merged_all.loc[index, 'L3 - HPSM'] == 'Áno':
                df_merged_all.loc[index, 'Group'] = 'Tollnet L3'
            else:
                df_merged_all.loc[index, 'Group'] = 'Tollnet'

    for index, value in df_merged_all['P'].items():
        if np.isnan(value):
            df_merged_all.loc[index, 'P'] = df_merged_all.loc[index, 'Priority - HPSM']

    for index, value in df_merged_all['Assign Time'].items():
        if pd.isnull(value):  # kontrola ci je NaT
            df_merged_all.loc[index, 'Assign Time'] = df_merged_all.loc[index, 'SLT Start time - HPSM']

    report_cols = ['Čas parametra S.2', 'Splnenie parametra S.2',
                   'Čas parametra S.3', 'Splnenie parametra S.3',
                   'Čas parametra S.4', 'Splnenie parametra S.4',
                   'Čas parametra S.5', 'Splnenie parametra S.5',
                   'Čas parametra S.6', 'Splnenie parametra S.6']
    report_cols_extra = ['L2 Odozva HPSM', 'L2 Odozva Breach HPSM',  # S2
                         'L2 Riesenie HPSM', 'L2 Riesenie Breach HPSM',  # S3
                         'L3 Odozva HPSM', 'L3 Odozva Breach HPSM',  # S4
                         'L3 Odozva HPSM', 'L3 Odozva Breach HPSM',  # S5 - rovnake ako S4
                         'L3 Riesenie HPSM', 'L3 Riesenie Breach HPSM']  # S6
    breached_hpsm_cols = ['L2 Odozva Breach HPSM', 'L2 Riesenie Breach HPSM',
                          'L3 Odozva Breach HPSM', 'L3 Odozva Breach HPSM',
                          'L3 Riesenie Breach HPSM']

    # prehodenie logiky HPSM Breached
    for index_breached, value_breached in enumerate(breached_hpsm_cols):
        for index, value in df_merged_all[value_breached].items():
            if value == 'Nie':
                df_merged_all.loc[index, value_breached] = 'Áno'
            elif value == 'Áno':
                df_merged_all.loc[index, value_breached] = 'Nie'

    for index_rep_col, value_rep_col in enumerate(report_cols):  # iterovanie cez zoznamy hore
        for index, value in df_merged_all[value_rep_col].items():
            if value is np.nan:
                df_merged_all.loc[index, value_rep_col] = df_merged_all.loc[index, report_cols_extra[index_rep_col]]

    # 3.8 - vyhodenie uzavretych incidentov z minulych obdobi
    for index, value in df_merged_all['Status JIRA'].items():  # data z interneho minulomesacneho reportu
        if value == 'Closed' or (value is np.nan and df_merged_all.loc[index, 'Status - JIRA'] == 'Closed'):
            # jednoduchy Closed ked uzatvoreny v minulom obdobi
            # incident v exporte zo starsieho obdobia = bez stavu reporte, JIRA vsak detekuje uzavretie
            df_merged_all.drop(labels=index, inplace=True, axis=0)

    # 3.9 - update JIRA stavov -> kontrola ci je co updatenut, potom update
    for index, value in df_merged_all['Status - JIRA'].items():
        if value is not np.nan:
            # skratenie dlhych nazvov stavov
            if value == 'Waiting for customer':
                df_merged_all.loc[index, 'Status - JIRA'] = 'WFC'
            elif value == 'Ready For Test':
                df_merged_all.loc[index, 'Status - JIRA'] = 'RFT'
            # update stavu
            df_merged_all.loc[index, 'Status JIRA'] = df_merged_all.loc[index, 'Status - JIRA']

    # 3.10.1 - vytvorenie zoznamu sviatkov z excelu
    sviatky_list = []
    workbook_sviatky = opx.load_workbook(filename='SVIATKY_DO_2022.xlsx')
    sheet_sviatky = workbook_sviatky.active
    row_sviatky = 1

    while True:
        if sheet_sviatky.cell(row=row_sviatky, column=1).value:
            sviatky_list.append(dt.datetime.strptime(sheet_sviatky.cell(row=row_sviatky, column=1).value, '%d.%m.%Y'))
            row_sviatky += 1
        else:
            break

    workbook_sviatky.close()

    # TODO - SKONTROLOVAT CI POCITA SPRAVNE
    # 3.10.2 - vyznacenie uz vyhodnotenych incidentov z minulych mesiacov + doplnenie prazdnych casov
    for index_rep_col, value_rep_col in enumerate(report_cols[::2]):  # iterovanie cez zoznamy hore
        for index, value in df_merged_all[value_rep_col].items():
            if value != 'už vyhodnotené' and value:
                if value.month < int(report_month) and value.year < int(report_year) and \
                        ((value_rep_col != 'Čas parametra S.3' and df_merged_all.loc[index, 'Group'] == 'Tollnet') or
                         (value_rep_col != 'Čas parametra S.6' and df_merged_all.loc[index, 'Group'] == 'Tollnet L3')):
                    df_merged_all.loc[index, value_rep_col] = 'už vyhodnotené'
                    df_merged_all.loc[index, report_cols[index_rep_col]] = 'už vyhodnotené'

                # doplnenie prazdnych casov na zaklade SLA dat z excelu
                if not value:
                    if value_rep_col == 'Čas parametra S.2':
                        # najdenie hodnoty a jednotky casu
                        if df_merged_all.loc[index, 'Group'] == 'Tollnet':
                            time_add = aux_functions.get_TOIS_SLA(df_merged_all.loc[index, 'P'], 'L2O')
                        else:
                            time_add = aux_functions.get_TOIS_SLA(df_merged_all.loc[index, 'P'], 'L3O')

                        # predefinovanie do formatu datetime
                        assign_time = df_merged_all.loc[index, 'Assign Time']
                        if time_add[1] == 'min':
                            df_merged_all.loc[index, value_rep_col] = assign_time + relativedelta(minutes=time_add[0])
                        elif time_add[1] == 'h':
                            df_merged_all.loc[index, value_rep_col] = assign_time + relativedelta(hours=time_add[0])
                        elif time_add[1] == 'day':
                            df_merged_all.loc[index, value_rep_col] = assign_time + relativedelta(days=time_add[0])
                        elif time_add[1] == 'workday':
                            df_merged_all.loc[index, value_rep_col] = wd.workday(assign_time, time_add[0], sviatky_list)
                        elif time_add[1] == 'BD':
                            pass
                            # manualne nastavene business day ako workday vzdy do 17:00
                            new_work_day = wd.workday(start_date=assign_time, days=time_add[0])
                            new_business_day = new_work_day.replace(hour=17, minute=0, second=0)
                            df_merged_all.loc[index, value_rep_col] = new_business_day
                        elif time_add[1] == 'month':
                            pass
                            df_merged_all.loc[index, value_rep_col] = assign_time + relativedelta(months=time_add[0])

                        # default SLA splnene
                        df_merged_all.loc[index, report_cols[index_rep_col]] = 'Áno'

    # 4.0 - ulozit do .xlsx pre rychly pristup + kontrolu, predtym konverzia na casovy format excelu
    excel_report_name = 'TOIS_report_created_' + dt.datetime.now().strftime("%Y%m%d_%H_%M_%S") + '.xlsx'
    excel_report_name_short = 'TOIS_report_created_short_' + dt.datetime.now().strftime("%Y%m%d_%H_%M_%S") + '.xlsx'
    df_merged_all.to_excel(excel_report_name, index=False, sheet_name=sht_name, freeze_panes=(1, 1))
    print("Report JIRA + HPSM pre TOIS uspesne ulozeny ako " + excel_report_name + ' ...')
    if to_gui:
        window['MLINE'].update("Report JIRA + HPSM pre TOIS uspesne ulozeny ako " + excel_report_name + ' ...\n', append=True)

    # 4.1 - vytvorit skrateny excel len so stlpcami ako v internom reporte
    cols_extra = ['Issue key - JIRA', 'Issue ID - JIRA', 'L3 - HPSM', 'Label 1 - JIRA', 'Label 2 - JIRA',
                  'Label 3 - JIRA',
                  'Priority - HPSM', 'Priority - JIRA', 'Status - HPSM', 'Status - JIRA', 'Resolution - JIRA',
                  'Assignee - JIRA', 'Assignee HPSM - JIRA', 'Created - JIRA', 'SLT Start time - HPSM',
                  'L2 Odozva HPSM', 'L2 Odozva HPSM Total Time', 'L2 Odozva Breach HPSM',
                  'L2 Riesenie HPSM', 'L2 Riesenie HPSM Total Time', 'L2 Riesenie Breach HPSM',
                  'L3 Odozva HPSM', 'L3 Odozva HPSM Total Time', 'L3 Odozva Breach HPSM',
                  'L3 Riesenie HPSM', 'L3 Riesenie HPSM Total Time', 'L3 Riesenie Breach HPSM',
                  'Updated - JIRA', 'Last Viewed - JIRA', 'Resolved - JIRA', 'Description - JIRA',
                  'Duplicate - JIRA', 'Relation to - JIRA', 'Outage - HPSM', 'HPSM Group - JIRA',
                  'HPSM Issue Type - JIRA', 'MEV ID - JIRA', 'Module - JIRA', 'Open-Closed Time - JIRA',
                  'Open-Resolved Time - JIRA', 'Reopen Counter - JIRA', 'Test Environment - JIRA']
    df_merged_all_short = df_merged_all.drop(cols_extra, axis=1, inplace=True)
    df_merged_all.to_excel(excel_report_name_short, index=False, sheet_name=sht_name, freeze_panes=(1, 1))
    print("Skrateny report JIRA + HPSM pre TOIS uspesne ulozeny ako " + excel_report_name + ' ...')
    if to_gui:
        window['MLINE'].update("Skrateny report JIRA + HPSM pre TOIS uspesne ulozeny ako " + excel_report_name + ' ...\n', append=True)

    # 5.0 - spracovanie vytvoreneho .xlsx spojeneho reportu // opx = openpyxl -> nastroj na pracovanie s xlsx
    workbook = opx.load_workbook(filename=excel_report_name)
    workbook_short = opx.load_workbook(filename=excel_report_name_short)
    sheet = workbook.active
    sheet_short = workbook_short.active

    # 5.1 - nastavenie auto filtra + autofit sirky buniek
    sheet.auto_filter.ref = "A:BE"
    aux_functions.auto_format_cell_width(sheet)

    sheet_short.auto_filter.ref = "A:P"
    aux_functions.auto_format_cell_width(sheet_short)

    # 5.2 - skryt pomocne stlpce??
    if hidden_cols:  # nastavenie na zaciatku skriptu
        sheet.column_dimensions.group(start='B', end='C', hidden=True)
        sheet.column_dimensions.group(start='F', end='I', hidden=True)
        sheet.column_dimensions.group(start='K', end='L', hidden=True)
        sheet.column_dimensions.group(start='N', end='S', hidden=True)
        sheet.column_dimensions.group(start='U', end='W', hidden=True)
        sheet.column_dimensions['Y'].hidden = True
        sheet.column_dimensions['AB'].hidden = True
        sheet.column_dimensions['AE'].hidden = True
        sheet.column_dimensions['AJ'].hidden = True
        sheet.column_dimensions.group(start='AL', end='BF', hidden=True)

    # 5.X - ulozit subor
    workbook.save(filename=excel_report_name)
    workbook_short.save(filename=excel_report_name_short)
    print('Zakladne upravy v spojenom reporte uspesne zrealizovane ...')
    if to_gui:
        window['MLINE'].update('Zakladne upravy v spojenom reporte uspesne zrealizovane ...\n', append=True)

    # TODO - filter na zaklade zaciatocneho eDZ - case insensitive?
    # print(str(df_csv['Summary']).startswith('eDZ'))

    return excel_report_name, excel_report_name_short


# VOLANIE GUI
sg.change_look_and_feel('Dark Blue 3')  # please make your windows colorful

layout = [
    [sg.FileBrowse('Load Report', size=(10, 1), file_types=(("Excel Files", ".xlsx"),)), sg.Input(key='path_report')],
    [sg.Button('Sheet Name', size=(10, 1)), sg.InputText(key='sheet_report')],
    [sg.FileBrowse('Load JIRA', size=(10, 1), file_types=(("CSV Files", ".csv"),)), sg.Input(key='path_JIRA')],
    [sg.FileBrowse('Load HPSM', size=(10, 1), file_types=(("Excel Files", ".xlsx"),)), sg.Input(key='path_HPSM')],
    [sg.Text('Report Year', size=(10, 1)), sg.InputText(key='report_year', size=(5, 1)),
     sg.Text('Report Month', size=(10, 1)), sg.InputText(key='report_month', size=(5, 1)),
     sg.Button('RUN', size=(15, 1), button_color=('black', 'green'))],
    [sg.MLine(size=(60, 8), key='MLINE')]]

window = sg.Window('Merging reports', layout, size=(450, 300))

while True:  # Event Loop
    event, values = window.read()
    # print(event, values)
    if event == sg.WIN_CLOSED:
        break

    if event == 'Sheet Name':
        workbook_report = opx.load_workbook(filename=values['path_report'])
        window['sheet_report'].update(workbook_report.active.title)
        workbook_report.close()
        window['report_year'].update(workbook_report.active.title[-4:])
        window['report_month'].update(int(workbook_report.active.title[-6:-5]) + 1)

    if event == 'RUN':
        if values['path_report'] and values['sheet_report'] and values['path_JIRA']:
            if values['path_HPSM'] and values['report_year'] and values['report_month']:

                reports = merge_exports(report_path=values['path_report'],
                                        sht_name=values['sheet_report'],
                                        csv_input_path=values['path_JIRA'],
                                        hpsm_input_path=values['path_HPSM'],
                                        report_year=values['report_year'],
                                        report_month=values['report_month'])
                sg.popup_ok('Reporty uspesne vytvorene!', text_color='black', no_titlebar=True)
                window['MLINE'].update('Finalne reporty:\n', append=True)
                window['MLINE'].update(reports[0] + '\n', append=True)
                window['MLINE'].update(reports[1] + '\n', append=True)

            else:
                sg.popup_error('Nekompletne udaje - dopln a skus znova.',
                               background_color='orange', text_color='black', no_titlebar=True)
        else:
            sg.popup_error('Nekompletne udaje - dopln a skus znova.',
                           background_color='orange', text_color='black', no_titlebar=True)

window.close()
